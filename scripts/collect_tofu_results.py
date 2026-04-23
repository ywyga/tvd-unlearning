#!/usr/bin/env python3
"""
Collect TOFU evaluation results from all TOFU_SUMMARY.json files found
under a saves directory and write them to a single Excel file.

Usage:
    python scripts/collect_tofu_results.py [saves_dir] [--output results.xlsx]

The script finds every TOFU_SUMMARY.json recursively, infers the task name
from the path, and parses model / method / hyperparameters from it.

Recognised task-name patterns
------------------------------
  tofu_{model}_{forget_split}_TVD_lr{lr}_r{r}_d{d}_o{o}_n{n}
  tofu_{model}_{forget_split}_{method}_lr{lr}
  tofu_{model}_{forget_split}_{method}          (no lr in name)
  tofu_{model}_full[_{forget_split}]            (finetune baseline)
  tofu_{model}_{retain_split}                  (retain baseline)
  wmdp_{model}_{data_split}_TVD_lr{lr}_r{r}_d{d}_o{o}_n{n}_ms{N}
  wmdp_{model}_{data_split}_{method}_lr{lr}_ms{N}
  wmdp_{model}_{data_split}_TaskArithmetic_scale{s}
  wmdp_{model}_{data_split}_full               (finetune baseline)
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    sys.exit("pandas is required:  pip install pandas openpyxl")


# ── Ordered columns in the output sheet ───────────────────────────────────────

METADATA_COLS = [
    "task_name",
    "group",
    "checkpoint",
    "model",
    "forget_split",
    "data_split",
    "method",
    "lr",
    "scale",
    "max_samples",
    "lambda_reconstruct",
    "lambda_data",
    "lambda_orth",
    "lambda_norm",
    "path",
]

METRIC_COLS = [
    "forget_Truth_Ratio",
    "forget_Q_A_Prob",
    "forget_Q_A_ROUGE",
    "extraction_strength",
    "forget_quality",
    "privleak",
    "model_utility",
]

# Arrow appended to each metric column header.
# ↑ = higher is better,  ↓ = lower is better,  ≈0 = closer to 0 is better
METRIC_DIRECTION = {
    "forget_Truth_Ratio": "↑",
    "forget_Q_A_Prob":    "↓",
    "forget_Q_A_ROUGE":   "↓",
    "extraction_strength": "↓",
    "forget_quality":     "↑",
    "privleak":           "≈0",
    "model_utility":      "↑",
}

# path goes last so it doesn't clutter the view when comparing metrics
ALL_COLS = [c for c in METADATA_COLS if c != "path"] + METRIC_COLS + ["path"]

# ── Task-name parsing ──────────────────────────────────────────────────────────

# Forget splits are always forget01 / forget05 / forget10
_FORGET_RE = re.compile(r"_(forget\d+)_(.*)")

# TVD suffix:  TVD_lr{lr}_r{r}_d{d}_o{o}_n{n}
_TVD_RE = re.compile(
    r"TVD_lr([^_]+)_r([^_]+)_d([^_]+)_o([^_]+)_n([^_]+)$"
)

# TaskArithmetic suffix:  TaskArithmetic_scale{scale}
_TA_RE = re.compile(r"TaskArithmetic_scale(.+)$")

# Retain split names: retain99 / retain95 / retain90
_RETAIN_RE = re.compile(r"retain\d+$")

# WMDP data splits: cyber / bio / chem
_WMDP_SPLIT_RE = re.compile(r"_(cyber|bio|chem)_(.*)")

# Trailing _ms{N} suffix (max_samples encoded in task name)
_MS_RE = re.compile(r"_ms(\d+)$")


def _parse_method_part(method_part: str, info: dict) -> None:
    """Parse the method+hparams suffix into info dict (mutates in place)."""
    # Strip trailing _ms{N} (max_samples) before method parsing
    ms = _MS_RE.search(method_part)
    if ms:
        info["max_samples"] = int(ms.group(1))
        method_part = method_part[: ms.start()]

    # TVD?
    tvd = _TVD_RE.match(method_part)
    if tvd:
        info["method"] = "TVD"
        info["lr"] = tvd.group(1)
        info["lambda_reconstruct"] = _maybe_float(tvd.group(2))
        info["lambda_data"] = _maybe_float(tvd.group(3))
        info["lambda_orth"] = _maybe_float(tvd.group(4))
        info["lambda_norm"] = _maybe_float(tvd.group(5))
        return
    # TaskArithmetic?
    ta = _TA_RE.match(method_part)
    if ta:
        info["method"] = "TaskArithmetic"
        info["scale"] = _maybe_float(ta.group(1))
        return
    # Other method with lr
    other = re.match(r"([A-Za-z][A-Za-z0-9]*)_lr(.+)$", method_part)
    if other:
        info["method"] = other.group(1)
        info["lr"] = other.group(2)
        return
    # Method name only, no hyperparams encoded
    info["method"] = method_part


def parse_task_name(task_name: str) -> dict:
    """Return a dict of parsed fields from a task_name string."""
    info: dict = {k: None for k in METADATA_COLS}
    info["task_name"] = task_name

    # ── WMDP task names ────────────────────────────────────────────────────────
    if task_name.startswith("wmdp_"):
        body = task_name[len("wmdp_"):]
        m = _WMDP_SPLIT_RE.search(body)
        if m:
            info["model"] = body[: m.start()]
            info["data_split"] = m.group(1)
            method_part = m.group(2)
            # Finetune baselines: wmdp_{model}_{split}_full or _forget
            if method_part in ("full", "forget"):
                info["method"] = method_part
            else:
                _parse_method_part(method_part, info)
        else:
            info["model"] = body
        return info

    # ── TOFU task names ────────────────────────────────────────────────────────
    if not task_name.startswith("tofu_"):
        return info

    body = task_name[len("tofu_"):]

    # ── Does the name contain a forget split? ─────────────────────────────────
    m = _FORGET_RE.search(body)
    if m:
        info["model"] = body[: m.start()]
        info["forget_split"] = m.group(1)
        _parse_method_part(m.group(2), info)

    else:
        # No forget split → finetune baseline (full or retain)
        # body is e.g. "phi-1_5_full" or "phi-1_5_retain99"
        # The last token after the final underscore-group is the split name.
        # Model names may themselves contain underscores, so we check the suffix.
        if body.endswith("_full"):
            info["model"] = body[: -len("_full")]
            info["method"] = "full"
        else:
            rm = _RETAIN_RE.search(body)
            if rm:
                info["model"] = body[: rm.start() - 1]  # -1 for the underscore
                info["method"] = rm.group(0)             # e.g. "retain99"
            else:
                # Fallback: treat entire body as model name
                info["model"] = body

    return info


def _maybe_float(s: str):
    try:
        return float(s)
    except (ValueError, TypeError):
        return s


# ── Path → task_name + forget_split inference ─────────────────────────────────

_EVALS_FORGET_RE = re.compile(r"^evals_(forget\d+)$")
_FORGET_PART_RE = re.compile(r"^forget\d+$")
_CHECKPOINT_RE = re.compile(r"^checkpoint-(\d+)$")


def infer_from_path(summary_file: Path) -> tuple[str, str | None, str | None, str | None]:
    """
    Return (task_name, forget_split_override, group, checkpoint).

    task_name   — first path component starting with 'tofu_'
    forget_split_override — forget split read from an 'evals_forget*' dir when
                  not encoded in task_name (e.g. full-model evals)
    group       — first subdirectory under 'unlearn/' that is not itself a
                  task_name (i.e. does not start with 'tofu_'); None otherwise
    checkpoint  — checkpoint step number (str) when the path contains a
                  'checkpoint-{N}' component; None for final-model results
    """
    parts = summary_file.parts
    task_name: str | None = None
    forget_override: str | None = None
    group: str | None = None
    checkpoint: str | None = None

    # Find task_name
    for part in parts:
        if part.startswith("tofu_"):
            task_name = part
            break

    # Find group: the component immediately after 'unlearn' if it is not a
    # task_name itself
    for i, part in enumerate(parts):
        if part == "unlearn" and i + 1 < len(parts):
            candidate = parts[i + 1]
            if not candidate.startswith("tofu_"):
                group = candidate
            break

    # Find checkpoint
    for part in parts:
        m = _CHECKPOINT_RE.match(part)
        if m:
            checkpoint = m.group(1)
            break

    # Detect forget split from an evals_{split} or bare forget\d+ directory
    for part in parts:
        m = _EVALS_FORGET_RE.match(part)
        if m:
            forget_override = m.group(1)
            break
        if _FORGET_PART_RE.match(part):
            forget_override = part
            break

    if task_name is None:
        # Last-resort fallback
        parent = summary_file.parent
        task_name = parent.parent.name if parent.name.startswith("evals") else parent.name

    return task_name, forget_override, group, checkpoint


# ── Main collection ────────────────────────────────────────────────────────────

def collect_results(saves_dir: Path) -> list[dict]:
    rows = []
    for summary_file in sorted(saves_dir.rglob("TOFU_SUMMARY.json")):
        try:
            with open(summary_file, encoding="utf-8") as fh:
                summary = json.load(fh)
        except Exception as exc:
            print(f"Warning: could not read {summary_file}: {exc}", file=sys.stderr)
            continue

        task_name, forget_override, group, checkpoint = infer_from_path(summary_file)
        row = parse_task_name(task_name)

        # For full/retain baselines the forget split is in the directory, not
        # the task name — fill it in so the row can be joined with unlearn rows.
        if forget_override and row["forget_split"] is None:
            row["forget_split"] = forget_override

        row["group"] = group
        row["checkpoint"] = checkpoint
        row["path"] = str(summary_file.relative_to(saves_dir))

        for metric in METRIC_COLS:
            row[metric] = summary.get(metric, None)

        rows.append(row)

    return rows


# ── Excel output ───────────────────────────────────────────────────────────────

def write_excel(rows: list[dict], output: Path) -> None:
    df = pd.DataFrame(rows, columns=ALL_COLS)

    # Sort: model → forget_split → data_split → method, putting None last
    df = df.sort_values(
        ["model", "forget_split", "data_split", "method"],
        key=lambda col: col.fillna("\xff"),  # sorts None/NaN after real strings
        na_position="last",
    )

    # Rename metric columns to include direction arrows in the header
    df = df.rename(
        columns={
            col: f"{col} {METRIC_DIRECTION[col]}"
            for col in METRIC_COLS
            if col in METRIC_DIRECTION
        }
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")

        ws = writer.sheets["Results"]

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-size columns (cap at 40 chars)
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 2, 40
            )

        # Light header fill
        try:
            from openpyxl.styles import PatternFill, Font

            header_fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = bold
        except ImportError:
            pass  # openpyxl styles not critical


# ── CLI ────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Collect TOFU results into a single Excel file."
    )
    parser.add_argument(
        "saves_dir",
        nargs="?",
        default="saves",
        help="Root saves directory to search (default: saves)",
    )
    parser.add_argument(
        "--output",
        "-o",
        default="tofu_results.xlsx",
        help="Output Excel file path (default: tofu_results.xlsx)",
    )
    args = parser.parse_args()

    saves_dir = Path(args.saves_dir)
    if not saves_dir.exists():
        sys.exit(f"Error: directory not found: {saves_dir}")

    rows = collect_results(saves_dir)
    if not rows:
        sys.exit(f"No TOFU_SUMMARY.json files found under {saves_dir}")

    output = Path(args.output)
    write_excel(rows, output)
    print(f"Wrote {len(rows)} row(s) → {output}")


if __name__ == "__main__":
    main()
