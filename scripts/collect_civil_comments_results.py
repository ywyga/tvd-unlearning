#!/usr/bin/env python3
"""
Collect Civil Comments evaluation results from Detoxify_SUMMARY.json and
LMEval_SUMMARY.json files found under a saves directory and write them to
a single Excel file.

Usage:
    python scripts/collect_civil_comments_results.py [saves_dir] [--output results.xlsx]

The script finds every eval directory that contains at least one of
Detoxify_SUMMARY.json or LMEval_SUMMARY.json, merges their metrics,
infers the task name from the path, and parses model / method /
hyperparameters from it.

Recognised task-name patterns
------------------------------
  civil_comments_{model}_TVD_lr{lr}_r{r}_d{d}_o{o}_n{n}[_ms{N}]
  civil_comments_{model}_{method}_lr{lr}[_ms{N}]
  civil_comments_{model}_TaskArithmetic_scale{s}[_ms{N}]
  civil_comments_{model}_full                              (M1 finetune baseline)
  civil_comments_{model}_forget                            (M_forget_full baseline)
  civil_comments_{model}_forget_subset_{N}                 (M_forget_subset baseline)
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    sys.exit("pandas is required:  pip install pandas openpyxl")


# ── Ordered columns in the output sheet ───────────────────────────────────────

METADATA_COLS = [
    "task_name",
    "group",
    "checkpoint",
    "model",
    "method",
    "lr",
    "scale",
    "max_samples",
    "lambda_reconstruct",
    "lambda_data",
    "lambda_orth",
    "lambda_norm",
    "path",
]

METRIC_COLS = [
    "detoxify/mean_toxicity",
    "detoxify/toxicity_rate",
    "wikitext/word_perplexity",
]

METRIC_DIRECTION = {
    "detoxify/mean_toxicity":  "↓",
    "detoxify/toxicity_rate":  "↓",
    "wikitext/word_perplexity": "≈",
}

ALL_COLS = [c for c in METADATA_COLS if c != "path"] + METRIC_COLS + ["path"]


# ── Task-name parsing ──────────────────────────────────────────────────────────

# Matches known unlearning method names (capitalized)
_UPPER_METHOD_RE = re.compile(
    r"_(TVD|TaskArithmetic|GradAscent|GradDiff|NPO|RMU|SimNPO|DPO|AltPO|CEU|PDU|WGA|SatImp|UNDIAL)(_|$)"
)

# TVD suffix:  TVD_lr{lr}_r{r}_d{d}_o{o}_n{n}
_TVD_RE = re.compile(
    r"TVD_lr([^_]+)_r([^_]+)_d([^_]+)_o([^_]+)_n([^_]+)$"
)

# TaskArithmetic suffix:  TaskArithmetic_scale{scale}
_TA_RE = re.compile(r"TaskArithmetic_scale(.+)$")

# Trailing _ms{N} suffix (max_samples encoded in task name)
_MS_RE = re.compile(r"_ms(\d+)$")

# forget_subset_N suffix
_FORGET_SUBSET_RE = re.compile(r"_forget_subset_(\d+)$")

_CHECKPOINT_RE = re.compile(r"^checkpoint-(\d+)$")


def _maybe_float(s: str):
    try:
        return float(s)
    except (ValueError, TypeError):
        return s


def _parse_method_part(method_part: str, info: dict) -> None:
    """Parse the method+hparams suffix into info dict (mutates in place)."""
    ms = _MS_RE.search(method_part)
    if ms:
        info["max_samples"] = int(ms.group(1))
        method_part = method_part[: ms.start()]

    tvd = _TVD_RE.match(method_part)
    if tvd:
        info["method"] = "TVD"
        info["lr"] = tvd.group(1)
        info["lambda_reconstruct"] = _maybe_float(tvd.group(2))
        info["lambda_data"] = _maybe_float(tvd.group(3))
        info["lambda_orth"] = _maybe_float(tvd.group(4))
        info["lambda_norm"] = _maybe_float(tvd.group(5))
        return

    ta = _TA_RE.match(method_part)
    if ta:
        info["method"] = "TaskArithmetic"
        info["scale"] = _maybe_float(ta.group(1))
        return

    other = re.match(r"([A-Za-z][A-Za-z0-9]*)_lr(.+)$", method_part)
    if other:
        info["method"] = other.group(1)
        info["lr"] = other.group(2)
        return

    info["method"] = method_part


def parse_task_name(task_name: str) -> dict:
    """Return a dict of parsed fields from a civil_comments task_name string."""
    info: dict = {k: None for k in METADATA_COLS}
    info["task_name"] = task_name

    if not task_name.startswith("civil_comments_"):
        return info

    body = task_name[len("civil_comments_"):]

    # Finetune baseline: civil_comments_{model}_full
    if body.endswith("_full"):
        info["model"] = body[: -len("_full")]
        info["method"] = "full"
        return info

    # Finetune baseline: civil_comments_{model}_forget_subset_{N}
    fs = _FORGET_SUBSET_RE.search(body)
    if fs:
        info["model"] = body[: fs.start()]
        info["method"] = "forget_subset"
        info["max_samples"] = int(fs.group(1))
        return info

    # Finetune baseline: civil_comments_{model}_forget
    if body.endswith("_forget"):
        info["model"] = body[: -len("_forget")]
        info["method"] = "forget"
        return info

    # Unlearning methods (capitalized): split on first known method token
    m = _UPPER_METHOD_RE.search(body)
    if m:
        info["model"] = body[: m.start()]
        _parse_method_part(body[m.start() + 1:], info)  # +1 to skip leading _
        return info

    # Fallback: treat entire body as model name
    info["model"] = body
    return info


# ── Path → task_name + group + checkpoint inference ───────────────────────────

def infer_from_path(eval_dir: Path) -> tuple[str, str | None, str | None]:
    """
    Return (task_name, group, checkpoint) from an eval directory path.

    task_name  — the path component starting with 'civil_comments_'
    group      — subdirectory under 'unlearn/' that is not a task_name
    checkpoint — step number when a 'checkpoint-{N}' component is present
    """
    parts = eval_dir.parts
    task_name: str | None = None
    group: str | None = None
    checkpoint: str | None = None

    for part in parts:
        if part.startswith("civil_comments_"):
            task_name = part
            break

    for i, part in enumerate(parts):
        if part in ("unlearn", "finetune") and i + 1 < len(parts):
            candidate = parts[i + 1]
            if not candidate.startswith("civil_comments_"):
                group = candidate
            break

    for part in parts:
        m = _CHECKPOINT_RE.match(part)
        if m:
            checkpoint = m.group(1)
            break

    if task_name is None:
        task_name = eval_dir.parent.name if eval_dir.name.startswith("evals") else eval_dir.name

    return task_name, group, checkpoint


# ── Main collection ────────────────────────────────────────────────────────────

_SUMMARY_FILES = ("Detoxify_SUMMARY.json", "LMEval_SUMMARY.json")


def collect_results(saves_dir: Path) -> list[dict]:
    # Find all eval directories that contain at least one known summary file.
    eval_dirs: set[Path] = set()
    for name in _SUMMARY_FILES:
        for f in saves_dir.rglob(name):
            eval_dirs.add(f.parent)

    rows = []
    for eval_dir in sorted(eval_dirs):
        # Merge metrics from all summary files present in this directory.
        merged: dict = {}
        for name in _SUMMARY_FILES:
            summary_file = eval_dir / name
            if not summary_file.exists():
                continue
            try:
                with open(summary_file, encoding="utf-8") as fh:
                    merged.update(json.load(fh))
            except Exception as exc:
                print(f"Warning: could not read {summary_file}: {exc}", file=sys.stderr)

        if not merged:
            continue

        task_name, group, checkpoint = infer_from_path(eval_dir)
        row = parse_task_name(task_name)
        row["group"] = group
        row["checkpoint"] = checkpoint
        row["path"] = str(eval_dir.relative_to(saves_dir))

        for metric in METRIC_COLS:
            row[metric] = merged.get(metric, None)

        rows.append(row)

    return rows


# ── Excel output ───────────────────────────────────────────────────────────────

def write_excel(rows: list[dict], output: Path) -> None:
    df = pd.DataFrame(rows, columns=ALL_COLS)

    df = df.sort_values(
        ["model", "method", "lr"],
        key=lambda col: col.fillna("\xff"),
        na_position="last",
    )

    df = df.rename(
        columns={
            col: f"{col} {METRIC_DIRECTION[col]}"
            for col in METRIC_COLS
            if col in METRIC_DIRECTION
        }
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")

        ws = writer.sheets["Results"]
        ws.freeze_panes = "A2"

        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

        try:
            from openpyxl.styles import PatternFill, Font

            header_fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = bold
        except ImportError:
            pass


# ── CLI ────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Collect Civil Comments unlearning results into a single Excel file."
    )
    parser.add_argument(
        "saves_dir",
        nargs="?",
        default="saves",
        help="Root saves directory to search (default: saves)",
    )
    parser.add_argument(
        "--output", "-o",
        default="civil_comments_results.xlsx",
        help="Output Excel file path (default: civil_comments_results.xlsx)",
    )
    args = parser.parse_args()

    saves_dir = Path(args.saves_dir)
    if not saves_dir.exists():
        sys.exit(f"Error: directory not found: {saves_dir}")

    rows = collect_results(saves_dir)
    if not rows:
        sys.exit(
            f"No Detoxify_SUMMARY.json or LMEval_SUMMARY.json files found under {saves_dir}"
        )

    output = Path(args.output)
    write_excel(rows, output)
    print(f"Wrote {len(rows)} row(s) → {output}")


if __name__ == "__main__":
    main()
